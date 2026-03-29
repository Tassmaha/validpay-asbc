"""
Core validation logic for ValidPay-ASBC.

Extracted from validapay.py to enable unit testing without Streamlit dependencies.
"""

import io

import pandas as pd
from openpyxl.styles import PatternFill


def normaliser_texte(valeur):
    """Normalize text: strip, uppercase, collapse whitespace."""
    return " ".join(str(valeur).strip().upper().split())


def nettoyer_telephone(valeur):
    """Remove non-digit characters from a phone number."""
    return "".join(ch for ch in str(valeur) if ch.isdigit())


def valider_format_tel(val):
    """Validate phone number format: must be exactly 8 digits."""
    val_str = str(val).strip()
    if not val_str.isdigit():
        return "Alphanumérique"
    if len(val_str) != 8:
        return "Longueur Incorrecte"
    return "OK"


def executer_validation(df_pay, df_ref, col_tel=None, cols_doublons=None, col_village=None):
    """
    Run the full validation pipeline on a payment DataFrame.

    Args:
        df_pay: Payment DataFrame (must have 'CLE_UNIQUE' column).
        df_ref: Reference DataFrame (must have 'CLE_UNIQUE' column).
        col_tel: Name of the phone column, or None to skip phone validation.
        cols_doublons: List of columns for duplicate detection, or empty list.
        col_village: Name of the village column, or None to skip village quota check.

    Returns:
        The df_pay DataFrame with a 'Statut_ValidaPay' column added/updated.
    """
    if cols_doublons is None:
        cols_doublons = []

    df_pay = df_pay.copy()
    df_pay["Statut_ValidaPay"] = "Valide"

    # 1. Phone format validation (lowest priority)
    if col_tel is not None and col_tel in df_pay.columns:
        verif_tel = df_pay[col_tel].apply(valider_format_tel)
        df_pay.loc[verif_tel != "OK", "Statut_ValidaPay"] = "Erreur Format Tel"

    # 2. Absent check (overrides phone error)
    mask_absent = ~df_pay["CLE_UNIQUE"].isin(df_ref["CLE_UNIQUE"])
    df_pay.loc[mask_absent, "Statut_ValidaPay"] = "Absent"

    # 3. Duplicate check (overrides absent)
    if cols_doublons:
        mask_doublon = df_pay.duplicated(subset=cols_doublons, keep=False)
        df_pay.loc[mask_doublon, "Statut_ValidaPay"] = "Doublon"

    # 4. Village quota (overrides all — max 2 ASBC per village)
    if col_village is not None and col_village in df_pay.columns:
        serie_village = df_pay[col_village].astype(str).str.strip()
        mask_village_valide = serie_village != ""
        comptes_village = serie_village[mask_village_valide].value_counts()
        villages_satures = comptes_village[comptes_village > 2].index
        mask_sur_effectif = serie_village.isin(villages_satures)
        df_pay.loc[mask_sur_effectif, "Statut_ValidaPay"] = "Quota Village Dépassé"

    return df_pay


def construire_contexte_ia(dataframe):
    """Build AI context string from validation results."""
    if dataframe is None or "Statut_ValidaPay" not in dataframe.columns:
        return "L'analyse n'est pas encore lancée."

    total_agents = len(dataframe)
    stats_statut = dataframe["Statut_ValidaPay"].value_counts().to_dict()
    taux_anomalies = (
        (1 - (stats_statut.get("Valide", 0) / total_agents)) * 100
        if total_agents > 0
        else 0
    )

    col_geo_assistant = next(
        (
            c
            for c in dataframe.columns
            if any(x in c.lower() for x in ["district", "ds", "région", "province"])
        ),
        None,
    )

    resume_geo = "Non disponible"
    if col_geo_assistant:
        erreurs_geo = dataframe[dataframe["Statut_ValidaPay"] != "Valide"]
        if not erreurs_geo.empty:
            top_geo = erreurs_geo[col_geo_assistant].value_counts().head(3).to_dict()
            resume_geo = f"Top zones en anomalies ({col_geo_assistant}): {top_geo}"

    return f"""
    Tu es l'analyste expert de la Direction de la Santé Communautaire au Burkina Faso.
    Réponds en français, de manière professionnelle et concise.
    Donne systématiquement :
    1) un constat chiffré,
    2) une interprétation,
    3) 2 recommandations opérationnelles.

    DONNÉES:
    - Total agents: {total_agents}
    - Répartition statuts: {stats_statut}
    - Taux anomalies: {taux_anomalies:.1f}%
    - {resume_geo}
    """


def reponse_assistant_local(dataframe, question):
    """Generate a local (non-API) analysis response."""
    if dataframe is None or "Statut_ValidaPay" not in dataframe.columns:
        return (
            "L'analyse n'est pas encore lancée. Chargez les fichiers, sélectionnez les colonnes clés "
            "puis relancez votre question."
        )

    total_agents = len(dataframe)
    stats_statut = dataframe["Statut_ValidaPay"].value_counts().to_dict()
    nb_valides = stats_statut.get("Valide", 0)
    nb_rejets = total_agents - nb_valides
    taux_anomalies = (nb_rejets / total_agents) * 100 if total_agents > 0 else 0

    top_anomalies = (
        dataframe[dataframe["Statut_ValidaPay"] != "Valide"]["Statut_ValidaPay"]
        .value_counts()
        .head(3)
        .to_dict()
    )

    col_geo_assistant = next(
        (
            c
            for c in dataframe.columns
            if any(x in c.lower() for x in ["district", "ds", "région", "province"])
        ),
        None,
    )
    detail_geo = "Non disponible"
    if col_geo_assistant:
        erreurs_geo = dataframe[dataframe["Statut_ValidaPay"] != "Valide"]
        if not erreurs_geo.empty:
            detail_geo = erreurs_geo[col_geo_assistant].value_counts().head(3).to_dict()

    return f"""
**Analyse locale (sans IA de gemini)**

1) **Constat chiffré**
- Total ASBC analysés: **{total_agents}**
- Valides: **{nb_valides}**
- Rejets: **{nb_rejets}**
- Taux d'anomalies: **{taux_anomalies:.1f}%**
- Principales anomalies: **{top_anomalies if top_anomalies else 'Aucune'}**

2) **Interprétation**
La question posée était: _{question}_.
Le niveau d'anomalies observé indique une priorité de correction sur les statuts les plus fréquents.
Top zones en anomalies: **{detail_geo}**.

3) **Recommandations**
- Corriger en priorité les statuts dominants puis relancer la validation.
- Mettre en place un contrôle en amont sur la saisie téléphone et la qualité des identifiants.
"""


def detecter_colonne_geo(columns):
    """Detect a geographic column by keyword matching on column names.

    Returns the first column name containing 'district', 'ds', 'région', or 'province'
    (case-insensitive), or None if no match.
    """
    return next(
        (c for c in columns if any(x in c.lower() for x in ["district", "ds", "région", "province"])),
        None,
    )


def generer_corrections(df_pay, colonnes_texte, col_tel=None):
    """Generate a preview DataFrame and a journal of proposed corrections.

    Args:
        df_pay: The payment DataFrame.
        colonnes_texte: List of text columns to normalize.
        col_tel: Phone column name, or None to skip phone cleaning.

    Returns:
        (df_preview, journal_corrections) where df_preview is the corrected copy
        and journal_corrections is a list of dicts describing each correction.
    """
    df_preview = df_pay.copy()
    journal_corrections = []

    for col in colonnes_texte:
        if col in df_preview.columns:
            serie_avant = df_preview[col].astype(str)
            serie_apres = serie_avant.apply(normaliser_texte)
            mask_modif = serie_avant != serie_apres
            if mask_modif.any():
                df_preview.loc[mask_modif, col] = serie_apres[mask_modif]
                for idx in df_preview.index[mask_modif]:
                    journal_corrections.append({
                        "Ligne": int(idx) + 2,
                        "Colonne": col,
                        "Ancienne valeur": serie_avant.loc[idx],
                        "Nouvelle valeur": serie_apres.loc[idx],
                        "Type correction": "Normalisation texte",
                    })

    if col_tel is not None and col_tel in df_preview.columns:
        serie_tel_avant = df_preview[col_tel].astype(str)
        serie_tel_nettoyee = serie_tel_avant.apply(nettoyer_telephone)
        mask_tel = (
            (serie_tel_avant != serie_tel_nettoyee)
            & (serie_tel_nettoyee.str.len() == 8)
            & (serie_tel_nettoyee.str.isdigit())
        )
        if mask_tel.any():
            df_preview.loc[mask_tel, col_tel] = serie_tel_nettoyee[mask_tel]
            for idx in df_preview.index[mask_tel]:
                journal_corrections.append({
                    "Ligne": int(idx) + 2,
                    "Colonne": col_tel,
                    "Ancienne valeur": serie_tel_avant.loc[idx],
                    "Nouvelle valeur": serie_tel_nettoyee.loc[idx],
                    "Type correction": "Nettoyage téléphone",
                })

    return df_preview, journal_corrections


def generer_rapport_colore(df_pay):
    """Generate a colored Excel report as bytes.

    Valid rows get green fill, invalid rows get red fill.
    The CLE_UNIQUE column is excluded from the export.

    Returns:
        bytes content of the .xlsx file.
    """
    buffer = io.BytesIO()
    df_export = df_pay.drop(columns=["CLE_UNIQUE"], errors="ignore")

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Validation")
        worksheet = writer.sheets["Validation"]

        vert_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        rouge_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        idx_statut = df_export.columns.get_loc("Statut_ValidaPay") + 1
        nb_cols = len(df_export.columns)

        for row_num in range(2, len(df_export) + 2):
            cell_statut = worksheet.cell(row=row_num, column=idx_statut).value
            val_statut = str(cell_statut).strip() if cell_statut else ""
            fill = vert_fill if val_statut == "Valide" else rouge_fill
            for col_num in range(1, nb_cols + 1):
                worksheet.cell(row=row_num, column=col_num).fill = fill

    return buffer.getvalue()


def generer_liste_valides(df_pay):
    """Generate an Excel file containing only valid ASBC entries.

    Returns:
        bytes content of the .xlsx file.
    """
    buffer = io.BytesIO()
    df_valides = df_pay[df_pay["Statut_ValidaPay"] == "Valide"].drop(
        columns=["CLE_UNIQUE", "Statut_ValidaPay"], errors="ignore"
    )
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_valides.to_excel(writer, index=False)
    return buffer.getvalue()


def generer_journal_corrections(journal_corrections):
    """Generate an Excel file for the correction journal.

    Args:
        journal_corrections: list of dicts with correction details.

    Returns:
        bytes content of the .xlsx file, or None if no corrections.
    """
    if not journal_corrections:
        return None
    buffer = io.BytesIO()
    df_journal = pd.DataFrame(journal_corrections)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_journal.to_excel(writer, index=False, sheet_name="Journal corrections")
    return buffer.getvalue()
