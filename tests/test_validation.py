"""
Tests for ValidPay-ASBC validation logic.

These tests cover the pure functions and core validation pipeline
extracted from validapay.py.
"""

import io

import openpyxl
import pandas as pd
import pytest

from validation import (
    construire_contexte_ia,
    detecter_colonne_geo,
    executer_validation,
    generer_corrections,
    generer_journal_corrections,
    generer_liste_valides,
    generer_rapport_colore,
    nettoyer_telephone,
    normaliser_texte,
    reponse_assistant_local,
    valider_format_tel,
)


# ─── normaliser_texte ────────────────────────────────────────────────────────


class TestNormaliserTexte:
    def test_basic_uppercase(self):
        assert normaliser_texte("jean dupont") == "JEAN DUPONT"

    def test_strips_leading_trailing_whitespace(self):
        assert normaliser_texte("  jean  ") == "JEAN"

    def test_collapses_internal_whitespace(self):
        assert normaliser_texte("jean   pierre   dupont") == "JEAN PIERRE DUPONT"

    def test_already_normalized(self):
        assert normaliser_texte("JEAN") == "JEAN"

    def test_empty_string(self):
        assert normaliser_texte("") == ""

    def test_numeric_input(self):
        assert normaliser_texte(123) == "123"

    def test_none_input(self):
        assert normaliser_texte(None) == "NONE"

    def test_nan_input(self):
        assert normaliser_texte(float("nan")) == "NAN"

    def test_tabs_and_newlines(self):
        result = normaliser_texte("jean\t\n dupont")
        assert result == "JEAN DUPONT"


# ─── nettoyer_telephone ─────────────────────────────────────────────────────


class TestNettoyerTelephone:
    def test_pure_digits_unchanged(self):
        assert nettoyer_telephone("70123456") == "70123456"

    def test_removes_letters(self):
        assert nettoyer_telephone("70aa5522") == "705522"

    def test_removes_spaces(self):
        assert nettoyer_telephone("70 12 34 56") == "70123456"

    def test_removes_dashes(self):
        assert nettoyer_telephone("70-12-34-56") == "70123456"

    def test_removes_dots(self):
        assert nettoyer_telephone("70.12.34.56") == "70123456"

    def test_removes_plus_prefix(self):
        assert nettoyer_telephone("+22670123456") == "22670123456"

    def test_empty_string(self):
        assert nettoyer_telephone("") == ""

    def test_no_digits(self):
        assert nettoyer_telephone("abcdef") == ""

    def test_float_string(self):
        assert nettoyer_telephone("70123456.0") == "701234560"

    def test_numeric_input(self):
        assert nettoyer_telephone(70123456) == "70123456"


# ─── valider_format_tel ─────────────────────────────────────────────────────


class TestValiderFormatTel:
    def test_valid_8_digits(self):
        assert valider_format_tel("70123456") == "OK"

    def test_alphanumeric(self):
        assert valider_format_tel("70aa5522") == "Alphanumérique"

    def test_too_short(self):
        assert valider_format_tel("7012") == "Longueur Incorrecte"

    def test_too_long(self):
        assert valider_format_tel("701234567890") == "Longueur Incorrecte"

    def test_empty_string(self):
        assert valider_format_tel("") == "Alphanumérique"

    def test_with_spaces_stripped(self):
        assert valider_format_tel("  70123456  ") == "OK"

    def test_letters_only(self):
        assert valider_format_tel("abcdefgh") == "Alphanumérique"

    def test_special_characters(self):
        assert valider_format_tel("70-12-34") == "Alphanumérique"


# ─── executer_validation ────────────────────────────────────────────────────


class TestExecuterValidation:
    @staticmethod
    def _make_ref(names):
        """Helper to create a reference DataFrame."""
        return pd.DataFrame({"Nom": names, "CLE_UNIQUE": names})

    @staticmethod
    def _make_pay(names, phones=None, villages=None):
        """Helper to create a payment DataFrame."""
        data = {"Nom": names, "CLE_UNIQUE": names}
        if phones is not None:
            data["Telephone"] = phones
        if villages is not None:
            data["Village"] = villages
        return pd.DataFrame(data)

    def test_all_valid(self):
        df_ref = self._make_ref(["ALICE", "BOB"])
        df_pay = self._make_pay(["ALICE", "BOB"])
        result = executer_validation(df_pay, df_ref, col_tel=None, cols_doublons=[], col_village=None)
        assert (result["Statut_ValidaPay"] == "Valide").all()

    def test_absent_detected(self):
        df_ref = self._make_ref(["ALICE"])
        df_pay = self._make_pay(["ALICE", "CHARLIE"])
        result = executer_validation(df_pay, df_ref, col_tel=None, cols_doublons=[], col_village=None)
        assert result.loc[result["Nom"] == "CHARLIE", "Statut_ValidaPay"].iloc[0] == "Absent"

    def test_duplicate_detected(self):
        df_ref = self._make_ref(["ALICE", "ALICE"])
        df_pay = self._make_pay(["ALICE", "ALICE"])
        result = executer_validation(df_pay, df_ref, col_tel=None, cols_doublons=["Nom"], col_village=None)
        assert (result["Statut_ValidaPay"] == "Doublon").all()

    def test_phone_error_detected(self):
        df_ref = self._make_ref(["ALICE"])
        df_pay = self._make_pay(["ALICE"], phones=["123"])
        result = executer_validation(df_pay, df_ref, col_tel="Telephone", cols_doublons=[], col_village=None)
        assert result["Statut_ValidaPay"].iloc[0] == "Erreur Format Tel"

    def test_valid_phone_no_error(self):
        df_ref = self._make_ref(["ALICE"])
        df_pay = self._make_pay(["ALICE"], phones=["70123456"])
        result = executer_validation(df_pay, df_ref, col_tel="Telephone", cols_doublons=[], col_village=None)
        assert result["Statut_ValidaPay"].iloc[0] == "Valide"

    def test_village_quota_exceeded(self):
        df_ref = self._make_ref(["A", "B", "C"])
        df_pay = self._make_pay(["A", "B", "C"], villages=["V1", "V1", "V1"])
        result = executer_validation(df_pay, df_ref, col_tel=None, cols_doublons=[], col_village="Village")
        assert (result["Statut_ValidaPay"] == "Quota Village Dépassé").all()

    def test_village_quota_within_limit(self):
        df_ref = self._make_ref(["A", "B"])
        df_pay = self._make_pay(["A", "B"], villages=["V1", "V1"])
        result = executer_validation(df_pay, df_ref, col_tel=None, cols_doublons=[], col_village="Village")
        assert (result["Statut_ValidaPay"] == "Valide").all()

    def test_priority_absent_over_phone_error(self):
        """Absent status should override phone format error."""
        df_ref = self._make_ref(["ALICE"])
        df_pay = self._make_pay(["BOB"], phones=["bad"])
        result = executer_validation(df_pay, df_ref, col_tel="Telephone", cols_doublons=[], col_village=None)
        assert result["Statut_ValidaPay"].iloc[0] == "Absent"

    def test_priority_doublon_over_absent(self):
        """Doublon status should override absent status."""
        df_ref = self._make_ref(["ALICE"])
        df_pay = self._make_pay(["BOB", "BOB"])
        result = executer_validation(df_pay, df_ref, col_tel=None, cols_doublons=["Nom"], col_village=None)
        assert (result["Statut_ValidaPay"] == "Doublon").all()

    def test_empty_payment_dataframe(self):
        df_ref = self._make_ref(["ALICE"])
        df_pay = self._make_pay([])
        result = executer_validation(df_pay, df_ref, col_tel=None, cols_doublons=[], col_village=None)
        assert len(result) == 0


# ─── construire_contexte_ia ─────────────────────────────────────────────────


class TestConstruireContexteIA:
    def test_none_dataframe(self):
        result = construire_contexte_ia(None)
        assert "pas encore lancée" in result

    def test_missing_column(self):
        df = pd.DataFrame({"col": [1, 2]})
        result = construire_contexte_ia(df)
        assert "pas encore lancée" in result

    def test_valid_dataframe(self):
        df = pd.DataFrame({"Statut_ValidaPay": ["Valide", "Absent", "Valide"]})
        result = construire_contexte_ia(df)
        assert "Total agents: 3" in result
        assert "33.3%" in result

    def test_geo_column_detected(self):
        df = pd.DataFrame({
            "Statut_ValidaPay": ["Valide", "Absent"],
            "District": ["D1", "D1"],
        })
        result = construire_contexte_ia(df)
        assert "District" in result


# ─── reponse_assistant_local ────────────────────────────────────────────────


class TestReponseAssistantLocal:
    def test_none_dataframe(self):
        result = reponse_assistant_local(None, "test question")
        assert "pas encore lancée" in result

    def test_missing_column(self):
        df = pd.DataFrame({"col": [1]})
        result = reponse_assistant_local(df, "test")
        assert "pas encore lancée" in result

    def test_includes_question(self):
        df = pd.DataFrame({"Statut_ValidaPay": ["Valide", "Absent"]})
        result = reponse_assistant_local(df, "Quel est le taux?")
        assert "Quel est le taux?" in result

    def test_correct_counts(self):
        df = pd.DataFrame({"Statut_ValidaPay": ["Valide", "Absent", "Doublon"]})
        result = reponse_assistant_local(df, "analyse")
        assert "**3**" in result  # total
        assert "**1**" in result  # valides
        assert "**2**" in result  # rejets

    def test_zero_anomalies(self):
        df = pd.DataFrame({"Statut_ValidaPay": ["Valide", "Valide"]})
        result = reponse_assistant_local(df, "analyse")
        assert "0.0%" in result

    def test_geo_column_in_response(self):
        df = pd.DataFrame({
            "Statut_ValidaPay": ["Valide", "Absent"],
            "District": ["D1", "D2"],
        })
        result = reponse_assistant_local(df, "analyse geo")
        assert "D2" in result


# ─── detecter_colonne_geo ───────────────────────────────────────────────────


class TestDetecterColonneGeo:
    def test_district_detected(self):
        assert detecter_colonne_geo(["Nom", "District Sanitaire", "Tel"]) == "District Sanitaire"

    def test_ds_detected(self):
        assert detecter_colonne_geo(["Nom", "DS", "Tel"]) == "DS"

    def test_region_detected(self):
        assert detecter_colonne_geo(["Nom", "Région", "Tel"]) == "Région"

    def test_province_detected(self):
        assert detecter_colonne_geo(["Nom", "Province", "Tel"]) == "Province"

    def test_case_insensitive(self):
        assert detecter_colonne_geo(["nom", "DISTRICT", "tel"]) == "DISTRICT"

    def test_no_match(self):
        assert detecter_colonne_geo(["Nom", "Prénom", "Tel"]) is None

    def test_empty_columns(self):
        assert detecter_colonne_geo([]) is None

    def test_first_match_returned(self):
        result = detecter_colonne_geo(["Province", "District"])
        assert result == "Province"


# ─── generer_corrections ────────────────────────────────────────────────────


class TestGenererCorrections:
    def test_text_normalization_detected(self):
        df = pd.DataFrame({"Nom": ["  jean  dupont  "], "CLE_UNIQUE": ["x"]})
        preview, journal = generer_corrections(df, ["Nom"])
        assert len(journal) == 1
        assert journal[0]["Type correction"] == "Normalisation texte"
        assert preview["Nom"].iloc[0] == "JEAN DUPONT"

    def test_no_correction_needed(self):
        df = pd.DataFrame({"Nom": ["JEAN"], "CLE_UNIQUE": ["x"]})
        preview, journal = generer_corrections(df, ["Nom"])
        assert len(journal) == 0

    def test_phone_cleaning_detected(self):
        df = pd.DataFrame({"Tel": ["70-12-34-56"], "CLE_UNIQUE": ["x"]})
        preview, journal = generer_corrections(df, [], col_tel="Tel")
        assert len(journal) == 1
        assert journal[0]["Type correction"] == "Nettoyage téléphone"
        assert preview["Tel"].iloc[0] == "70123456"

    def test_phone_not_cleaned_if_result_not_8_digits(self):
        """Phone with letters that produces <8 digits after cleaning should NOT be proposed."""
        df = pd.DataFrame({"Tel": ["70aa55"], "CLE_UNIQUE": ["x"]})
        preview, journal = generer_corrections(df, [], col_tel="Tel")
        phone_corrections = [j for j in journal if j["Type correction"] == "Nettoyage téléphone"]
        assert len(phone_corrections) == 0

    def test_multiple_columns_corrected(self):
        df = pd.DataFrame({
            "Nom": ["  alice  "],
            "Prenom": ["  bob  "],
            "CLE_UNIQUE": ["x"],
        })
        preview, journal = generer_corrections(df, ["Nom", "Prenom"])
        assert len(journal) == 2

    def test_missing_column_skipped(self):
        df = pd.DataFrame({"Nom": ["jean"], "CLE_UNIQUE": ["x"]})
        preview, journal = generer_corrections(df, ["Nom", "ColInexistante"])
        # Should not raise, just skip the missing column
        assert isinstance(journal, list)

    def test_journal_line_numbers(self):
        """Line numbers should be index + 2 (accounting for Excel header row)."""
        df = pd.DataFrame({"Nom": ["CLEAN", "  fix me  "], "CLE_UNIQUE": ["x", "y"]})
        _, journal = generer_corrections(df, ["Nom"])
        assert len(journal) == 1
        assert journal[0]["Ligne"] == 3  # index 1 + 2


# ─── generer_rapport_colore ─────────────────────────────────────────────────


class TestGenererRapportColore:
    def _make_df(self):
        return pd.DataFrame({
            "Nom": ["ALICE", "BOB"],
            "Statut_ValidaPay": ["Valide", "Absent"],
            "CLE_UNIQUE": ["ALICE", "BOB"],
        })

    def test_returns_valid_xlsx(self):
        data = generer_rapport_colore(self._make_df())
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.title == "Validation"

    def test_cle_unique_excluded(self):
        data = generer_rapport_colore(self._make_df())
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "CLE_UNIQUE" not in headers

    def test_valid_row_green_fill(self):
        data = generer_rapport_colore(self._make_df())
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        # Row 2 is ALICE (Valide) → green
        fill_color = ws.cell(row=2, column=1).fill.start_color.rgb
        assert fill_color == "00C6EFCE"

    def test_invalid_row_red_fill(self):
        data = generer_rapport_colore(self._make_df())
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        # Row 3 is BOB (Absent) → red
        fill_color = ws.cell(row=3, column=1).fill.start_color.rgb
        assert fill_color == "00FFC7CE"


# ─── generer_liste_valides ──────────────────────────────────────────────────


class TestGenererListeValides:
    def test_only_valid_rows(self):
        df = pd.DataFrame({
            "Nom": ["ALICE", "BOB", "CHARLIE"],
            "Statut_ValidaPay": ["Valide", "Absent", "Valide"],
            "CLE_UNIQUE": ["A", "B", "C"],
        })
        data = generer_liste_valides(df)
        result = pd.read_excel(io.BytesIO(data))
        assert len(result) == 2
        assert set(result["Nom"]) == {"ALICE", "CHARLIE"}

    def test_excludes_status_and_key_columns(self):
        df = pd.DataFrame({
            "Nom": ["ALICE"],
            "Statut_ValidaPay": ["Valide"],
            "CLE_UNIQUE": ["A"],
        })
        data = generer_liste_valides(df)
        result = pd.read_excel(io.BytesIO(data))
        assert "Statut_ValidaPay" not in result.columns
        assert "CLE_UNIQUE" not in result.columns

    def test_empty_when_all_rejected(self):
        df = pd.DataFrame({
            "Nom": ["BOB"],
            "Statut_ValidaPay": ["Absent"],
            "CLE_UNIQUE": ["B"],
        })
        data = generer_liste_valides(df)
        result = pd.read_excel(io.BytesIO(data))
        assert len(result) == 0


# ─── generer_journal_corrections ────────────────────────────────────────────


class TestGenererJournalCorrections:
    def test_returns_none_for_empty(self):
        assert generer_journal_corrections([]) is None

    def test_returns_valid_xlsx(self):
        journal = [{"Ligne": 2, "Colonne": "Nom", "Ancienne valeur": "jean", "Nouvelle valeur": "JEAN", "Type correction": "Normalisation texte"}]
        data = generer_journal_corrections(journal)
        assert data is not None
        result = pd.read_excel(io.BytesIO(data), sheet_name="Journal corrections")
        assert len(result) == 1
        assert result["Colonne"].iloc[0] == "Nom"

    def test_multiple_corrections(self):
        journal = [
            {"Ligne": 2, "Colonne": "Nom", "Ancienne valeur": "a", "Nouvelle valeur": "A", "Type correction": "Normalisation texte"},
            {"Ligne": 3, "Colonne": "Tel", "Ancienne valeur": "70-00-00-00", "Nouvelle valeur": "70000000", "Type correction": "Nettoyage téléphone"},
        ]
        data = generer_journal_corrections(journal)
        result = pd.read_excel(io.BytesIO(data), sheet_name="Journal corrections")
        assert len(result) == 2
