import streamlit as st
import pandas as pd
import io
import streamlit_folium as st_folium
import folium
import google.generativeai as genai
from PIL import Image
from openpyxl.styles import PatternFill

# Configuration de l'IA (Utilisez votre clé ici)
genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
model = genai.GenerativeModel('models/gemini-2.0-flash')

# Configuration de la page
st.set_page_config(page_title="ValidaPay Pro", page_icon="🇧🇫", layout="wide")

# 1. Affichage du Logo centré (avant le titre)
try:
    # On crée 3 colonnes pour centrer l'image dans celle du milieu
    col_l1, col_l2, col_l3 = st.columns([2, 1, 2])
    with col_l2:
        logo = Image.open("logo_sante.png") # Assurez-vous que le fichier est dans le dossier
        st.image(logo, use_container_width=True)
except:
    # Si l'image n'est pas trouvée, on affiche un espace discret
    st.write("")

# 2. Titre de l'application centré
st.markdown(
    """
    <h1 style='text-align: center; color: #007BFF; margin-top: -20px;'>
        ValidPay-ASBC
    </h1>
    <p style='text-align: center; font-size: 1.2em; font-weight: bold;'>
        Direction de la Santé Communautaire
    </p>
    """, 
    unsafe_allow_html=True
)

st.divider()

def normaliser_texte(valeur):
    return " ".join(str(valeur).strip().upper().split())


def nettoyer_telephone(valeur):
    return "".join(ch for ch in str(valeur) if ch.isdigit())

# 1. Chargement des fichiers
col1, col2 = st.columns(2)
with col1:
    st.subheader("1. Base de référence (PGA)")
    ref_file = st.file_uploader("Personnel de référence (Excel)", type=['xlsx'], key="ref")

with col2:
    st.subheader("2. Liste de paiement")
    pay_file = st.file_uploader("Liste à valider (Excel)", type=['xlsx'], key="pay")

if ref_file and pay_file:
    df_ref = pd.read_excel(ref_file, engine='calamine').astype(str)
    df_pay = pd.read_excel(pay_file, engine='calamine').astype(str)
    st.success("Fichiers chargés avec succès !")
    
    st.header("🔍 Configuration") 
    colonnes_pay = df_pay.columns.tolist()
    
    # --- AJOUT DU MONTANT DE L'INDEMNITÉ ---
    c_fin1, c_fin2 = st.columns([1, 2])
    with c_fin1:
        montant_indemnite = st.number_input("Indemnité par ASBC (FCFA)", value=20000, step=1000)
    with c_fin2:
        st.caption("Ce montant sera utilisé pour calculer l'impact financier des rejets (doublons, absents, etc.).")
    
    colonnes_pay = df_pay.columns.tolist()
    
    c_a, c_b, c_c = st.columns(3)
    c_a, c_b, c_c, c_d = st.columns(4)
    with c_a:
        cols_cles = st.multiselect("Colonnes clés ID (ex: Nom, Prénom) :", colonnes_pay)
    with c_b:
        cols_doublons = st.multiselect("Colonnes pour les doublons :", colonnes_pay)
    with c_c:
        col_tel = st.selectbox("Colonne Téléphone :", ["Aucune"] + colonnes_pay)
    with c_d:
        col_village = st.selectbox("Colonne Village :", ["Aucune"] + colonnes_pay)

    if cols_cles:
        # Création de la clé unique
        df_pay['CLE_UNIQUE'] = df_pay[cols_cles].agg('-'.join, axis=1)
        
        if all(c in df_ref.columns for c in cols_cles):
            df_ref['CLE_UNIQUE'] = df_ref[cols_cles].agg('-'.join, axis=1)
            
           
         # --- LOGIQUE DE VALIDATION CORRIGÉE ---
            
            # Initialisation du Statut
            df_pay['Statut_ValidaPay'] = 'Valide'

            # 1. Vérification du Téléphone (Format Strict & Alphanumérique)
            def executer_validation():
                # Initialisation du Statut
                df_pay['Statut_ValidaPay'] = 'Valide'

                # 1. Vérification du Téléphone (Format Strict & Alphanumérique)
                if col_tel != "Aucune":
                    def valider_format_tel(val):
                        val_str = str(val).strip()
                        # Détecte si le contenu n'est pas uniquement numérique (ex: "70aa5522")
                        if not val_str.isdigit():
                            return "Alphanumérique"
                        # Détecte si la longueur est différente de 8
                        if len(val_str) != 8:
                            return "Longueur Incorrecte"
                        return "OK"

                    # Application de la fonction
                    verif_tel = df_pay[col_tel].apply(valider_format_tel)

                    # Marquage des erreurs de téléphone
                    df_pay.loc[verif_tel != "OK", 'Statut_ValidaPay'] = 'Erreur Format Tel'

                # 2. Marquage des Absents (Prioritaire sur le format tel)
                mask_absent = ~df_pay['CLE_UNIQUE'].isin(df_ref['CLE_UNIQUE'])
                df_pay.loc[mask_absent, 'Statut_ValidaPay'] = 'Absent'

                # 3. Marquage des Doublons (Prioritaire sur le format tel)
                if cols_doublons:
                    mask_doublon = df_pay.duplicated(subset=cols_doublons, keep=False)
                    df_pay.loc[mask_doublon, 'Statut_ValidaPay'] = 'Doublon'

                # 4. Vérification du quota par village (max 2 ASBC)
                if col_village != "Aucune":
                    serie_village = df_pay[col_village].astype(str).str.strip()
                    mask_village_valide = serie_village != ""
                    comptes_village = serie_village[mask_village_valide].value_counts()
                    villages_satures = comptes_village[comptes_village > 2].index
                    mask_sur_effectif = serie_village.isin(villages_satures)
                    df_pay.loc[mask_sur_effectif, 'Statut_ValidaPay'] = 'Quota Village Dépassé'

            # Validation initiale
            executer_validation()

            # --- CORRECTION ASSISTÉE (V1) ---
            st.subheader("🛠️ Correction assistée")
            st.caption("Aperçu des corrections proposées (normalisation texte et nettoyage téléphone) avant recalcul des statuts.")

            colonnes_texte = sorted(set(cols_cles + cols_doublons + ([col_village] if col_village != "Aucune" else [])))
            df_preview = df_pay.copy()
            journal_corrections = []

            for col in colonnes_texte:
                if col in df_preview.columns:
                    serie_avant = df_preview[col].astype(str)
                    serie_apres = serie_avant.apply(normaliser_texte)
                    mask_modif = serie_avant != serie_apres
                    if mask_modif.any():
                        df_preview.loc[mask_modif, col] = serie_apres[mask_modif]
                        for idx in df_preview.index[mask_modif]:
                            journal_corrections.append({
                                "Ligne": int(idx) + 2,
                                "Colonne": col,
                                "Ancienne valeur": serie_avant.loc[idx],
                                "Nouvelle valeur": serie_apres.loc[idx],
                                "Type correction": "Normalisation texte"
                            })

            if col_tel != "Aucune":
                def valider_format_tel(val):
                    val_str = str(val).strip()
                    # Détecte si le contenu n'est pas uniquement numérique (ex: "70aa5522")
                    if not val_str.isdigit():
                        return "Alphanumérique"
                    # Détecte si la longueur est différente de 8
                    if len(val_str) != 8:
                        return "Longueur Incorrecte"
                    return "OK"
                
                # Application de la fonction
                verif_tel = df_pay[col_tel].apply(valider_format_tel)
                
                # Marquage des erreurs de téléphone
                df_pay.loc[verif_tel != "OK", 'Statut_ValidaPay'] = 'Erreur Format Tel'
                serie_tel_avant = df_preview[col_tel].astype(str)
                serie_tel_nettoyee = serie_tel_avant.apply(nettoyer_telephone)
                mask_tel = (
                    (serie_tel_avant != serie_tel_nettoyee)
                    & (serie_tel_nettoyee.str.len() == 8)
                    & (serie_tel_nettoyee.str.isdigit())
                )
                if mask_tel.any():
                    df_preview.loc[mask_tel, col_tel] = serie_tel_nettoyee[mask_tel]
                    for idx in df_preview.index[mask_tel]:
                        journal_corrections.append({
                            "Ligne": int(idx) + 2,
                            "Colonne": col_tel,
                            "Ancienne valeur": serie_tel_avant.loc[idx],
                            "Nouvelle valeur": serie_tel_nettoyee.loc[idx],
                            "Type correction": "Nettoyage téléphone"
                        })

            if journal_corrections:
                df_journal = pd.DataFrame(journal_corrections)
                st.info(f"{len(df_journal)} correction(s) potentielle(s) détectée(s).")
                st.dataframe(df_journal.head(100), use_container_width=True)

                if st.button("✅ Appliquer les corrections recommandées"):
                    df_pay = df_preview.copy()
                    df_pay['CLE_UNIQUE'] = df_pay[cols_cles].agg('-'.join, axis=1)
                    executer_validation()
                    st.success("Corrections appliquées et validation recalculée.")
            else:
                df_journal = pd.DataFrame()
                st.success("Aucune correction automatique suggérée.")


            # 2. Marquage des Absents (Prioritaire sur le format tel)
            mask_absent = ~df_pay['CLE_UNIQUE'].isin(df_ref['CLE_UNIQUE'])
            df_pay.loc[mask_absent, 'Statut_ValidaPay'] = 'Absent'
            
            # 3. Marquage des Doublons (Prioritaire sur le format tel)
            if cols_doublons:
                mask_doublon = df_pay.duplicated(subset=cols_doublons, keep=False)
                df_pay.loc[mask_doublon, 'Statut_ValidaPay'] = 'Doublon'
           
# --- AFFICHAGE DES RÉSULTATS (POSITIONNEMENT CORRIGÉ) ---
            st.divider()
            
            # Calcul des compteurs
            nb_valides = len(df_pay[df_pay['Statut_ValidaPay'] == 'Valide'])
            nb_absents = len(df_pay[df_pay['Statut_ValidaPay'] == 'Absent'])
            nb_doublons = len(df_pay[df_pay['Statut_ValidaPay'] == 'Doublon'])
            nb_village = len(df_pay[df_pay['Statut_ValidaPay'] == 'Quota Village Dépassé'])
            nb_tel = len(df_pay[df_pay['Statut_ValidaPay'] == 'Erreur Format Tel'])
            
            # Calcul financier
            nb_total_rejets = len(df_pay[df_pay['Statut_ValidaPay'] != 'Valide'])
            eco_totale = nb_total_rejets * montant_indemnite

            # Affichage des indicateurs en colonnes
            st.subheader("📊 Résultats de la validation")
            res = st.columns(5)
            res = st.columns(6)
            res[0].metric("✅ Valides", f"{nb_valides}")
            res[1].metric("❌ Absents", f"{nb_absents}")
            res[2].metric("👯 Doublons", f"{nb_doublons}")
            res[3].metric("📞 Erreurs Tel", f"{nb_tel}")
            res[4].metric("💰 Économie (FCFA)", f"{eco_totale:,.0f}")
            res[3].metric("🏘️ Quota Village", f"{nb_village}")
            res[4].metric("📞 Erreurs Tel", f"{nb_tel}")
            res[5].metric("💰 Économie (FCFA)", f"{eco_totale:,.0f}")

            # Message d'alerte si des erreurs existent
            if eco_totale > 0:
                st.info(f"💡 **Impact budgétaire :** Cette validation permet d'économiser **{eco_totale:,.0f} FCFA**.")

            # Tableau des erreurs
            st.subheader("📋 Détail des incohérences")
            anomalies = df_pay[df_pay['Statut_ValidaPay'] != 'Valide'].drop(columns=['CLE_UNIQUE'], errors='ignore')
            if not anomalies.empty:
                st.dataframe(anomalies, use_container_width=True)
            else:
                st.success("Aucune anomalie détectée sur cette liste !") 

            # --- GRAPHIQUE ANALYTIQUE PAR DISTRICT ---
            st.write("") 
            col_geo = next((c for c in df_pay.columns if any(x in c.lower() for x in ['district', 'ds', 'région', 'province'])), None)
            
            if col_geo:
                st.subheader(f"📊 Répartition des incohérences par {col_geo}")
                df_erreurs_graph = df_pay[df_pay['Statut_ValidaPay'] != 'Valide']
                
                if not df_erreurs_graph.empty:
                    chart_data = df_erreurs_graph.groupby(col_geo).size().reset_index(name='Nombre de rejets')
                    chart_data = chart_data.sort_values(by='Nombre de rejets', ascending=False)
                    st.bar_chart(data=chart_data, x=col_geo, y='Nombre de rejets', color="#FF4B4B")
                    st.bar_chart(chart_data.set_index(col_geo)['Nombre de rejets'])
                    st.caption(f"Zones où les erreurs de saisie, les doublons ou les absences sont les plus élevés.")
                else:
                    st.success("Aucune erreur à afficher sur le graphique !")
            
           
            # --- GÉNÉRATION DES EXPORTS ---
            st.divider()
            st.subheader("📥 Exportation des fichiers")
            
            # 1. Rapport Coloré
            buffer_complet = io.BytesIO()
            with pd.ExcelWriter(buffer_complet, engine='openpyxl') as writer:
                df_export = df_pay.drop(columns=['CLE_UNIQUE'])
                df_export.to_excel(writer, index=False, sheet_name='Validation')
                worksheet = writer.sheets['Validation']
                
                vert_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                rouge_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                idx_statut_excel = df_export.columns.get_loc('Statut_ValidaPay') + 1
                nb_cols = len(df_export.columns)
                
                for row_num in range(2, len(df_export) + 2):
                    cell_statut = worksheet.cell(row=row_num, column=idx_statut_excel).value
                    val_statut = str(cell_statut).strip() if cell_statut else ""
                    fill = vert_fill if val_statut == 'Valide' else rouge_fill
                    for col_num in range(1, nb_cols + 1):
                        worksheet.cell(row=row_num, column=col_num).fill = fill

            # 2. Liste Propre
            df_valides_only = df_pay[df_pay['Statut_ValidaPay'] == 'Valide'].drop(columns=['CLE_UNIQUE', 'Statut_ValidaPay'])
            buffer_valides = io.BytesIO()
            with pd.ExcelWriter(buffer_valides, engine='openpyxl') as writer:
                df_valides_only.to_excel(writer, index=False)

            # Boutons
            exp1, exp2 = st.columns(2)
            with exp1:
                st.download_button(label="📥 Télécharger le Rapport global avec les incohérences", data=buffer_complet.getvalue(), file_name="Rapport_ValidPay_Colore.xlsx")
            with exp2:
                st.download_button(label="📥 Télécharger la liste des ASBC validés", data=buffer_valides.getvalue(), file_name="Liste_ASBC_Valides.xlsx")

            if 'df_journal' in locals() and not df_journal.empty:
                buffer_journal = io.BytesIO()
                with pd.ExcelWriter(buffer_journal, engine='openpyxl') as writer:
                    df_journal.to_excel(writer, index=False, sheet_name='Journal corrections')
                st.download_button(
                    label="📥 Télécharger le journal des corrections",
                    data=buffer_journal.getvalue(),
                    file_name="Journal_Corrections_ValidPay.xlsx"
                )
                
        else:
            st.error("Les colonnes clés sélectionnées ne correspondent pas dans la base de référence.")
    else:
        st.info("Sélectionnez les colonnes clés pour lancer l'analyse.")

# --- SECTION CHATBOT ASSISTANT IA ---
st.divider()
st.header("🔬 Assistant IA")

if "messages" not in st.session_state:
    st.session_state.messages = []

for message in st.session_state.messages:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

if prompt := st.chat_input("Ex: Quel est le taux de rejet par district ?"):
    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user"):
        st.markdown(prompt)

    if 'df_pay' in locals() and 'Statut_ValidaPay' in df_pay.columns:
        total_a = len(df_pay)
        stats_a = df_pay['Statut_ValidaPay'].value_counts().to_dict()
